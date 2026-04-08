import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Side, Border
from .models import Podmiot, Obszar, Blok, Pytanie, GrupaNaglowek, GrupaPodmioty, AnkietaNaglowek, AnkietaPytania
import datetime
from django.db import models
import re


def get_data_by_grupanaglowek_id(grupanaglowek_id):
    # 1. Pobierz grupę i podmioty (bez zmian)
    grupa_naglowek = GrupaNaglowek.objects.get(id=grupanaglowek_id)
    grupa_podmioty = GrupaPodmioty.objects.filter(id_grupa=grupanaglowek_id).select_related('id_podmiotu')
    ankiety = AnkietaNaglowek.objects.filter(id_grupa=grupanaglowek_id)

    result = {
        'grupa_naglowek': {
            'id': grupa_naglowek.id,
            'nazwa': grupa_naglowek.nazwa,
        },
        'podmioty': [
            {'id': gp.id_podmiotu.id, 'kod': gp.id_podmiotu.kod, 'nazwa': gp.id_podmiotu.nazwa}
            for gp in grupa_podmioty
        ],
        'ankiety': [],
    }

    for ankieta in ankiety:
        # Pobieramy AnkietaPytania wraz z całą drabiną powiązań w górę
        ankieta_pytania = AnkietaPytania.objects.filter(id_ankieta_naglowek=ankieta.id).select_related(
            'id_podzapytania__id_pytania__id_podbloku__id_bloku__id_obszaru'
        )

        ankieta_data = {
            'id': ankieta.id,
            'nazwa': ankieta.nazwa,
            'elementy_ankiety': [],  # Tu będą spłaszczone dane o podzapytaniach
        }

        for ap in ankieta_pytania:

            pz = ap.id_podzapytania
            pyt = pz.id_pytania
            pb = pyt.id_podbloku
            bl = pb.id_bloku
            ob = bl.id_obszaru

            ankieta_data['elementy_ankiety'].append({
                'obszar': {'id': ob.id, 'kod': ob.kod, 'nazwa': ob.nazwa},
                'blok': {'id': bl.id, 'kod': bl.kod, 'tresc': bl.tresc},
                'podblok': {'id': pb.id, 'kod': pb.kod, 'tresc': pb.tresc},
                'pytanie': {'id': pyt.id, 'kod': pyt.kod, 'tresc': pyt.tresc},
                'podzapytanie': {
                    'id': pz.id,
                    'kod': pz.kod,
                    'tresc': pz.tresc,
                    'obligatoryjne': pz.obligatoryjne,
                    # Tutaj miejsce na dane, które będą wypełniane w ankiecie
                    'odpowiedz': None,
                    'uzasadnienie': ""
                }
            })

        result['ankiety'].append(ankieta_data)

    return result

    # Przykład użycia
    # grupanaglowek_id = 1  # Zastąp odpowiednim ID grupy nagłówkowej
    # data = get_data_by_grupanaglowek_id(grupanaglowek_id)
    # print(data)

dict_original_dict = {}


def transform_dict(original_dict):
    hierarchy = {}
    ankiety = original_dict.get('ankiety', [])

    for survey in ankiety:
        elementy = survey.get('elementy_ankiety', [])
        for item in elementy:
            # 1. Obszar
            ob = item['obszar']
            if ob['kod'] not in hierarchy:
                hierarchy[ob['kod']] = {'info': ob, 'bloki': {}}

            # 2. Blok
            bl = item['blok']
            curr_obs = hierarchy[ob['kod']]['bloki']
            if bl['kod'] not in curr_obs:
                curr_obs[bl['kod']] = {'info': bl, 'podbloki': {}}

            # 3. Podblok
            pb = item['podblok']
            curr_bl = curr_obs[bl['kod']]['podbloki']
            if pb['kod'] not in curr_bl:
                curr_bl[pb['kod']] = {'info': pb, 'pytania': {}}

            # 4. Pytanie
            pyt = item['pytanie']
            curr_pb = curr_bl[pb['kod']]['pytania']
            if pyt['kod'] not in curr_pb:
                curr_pb[pyt['kod']] = {'info': pyt, 'podzapytania': []}

            # 5. Podzapytanie
            curr_pb[pyt['kod']]['podzapytania'].append(item['podzapytanie'])

    # SORTOWANIE PODZAPYTAŃ wewnątrz każdego pytania
    for obs in hierarchy.values():
        for bl in obs['bloki'].values():
            for pb in bl['podbloki'].values():
                for pyt in pb['pytania'].values():
                    pyt['podzapytania'].sort(key=lambda x: natural_sort_key(x['kod']))

    return {
        'hierarchia': hierarchy,
        'podmioty': original_dict.get('podmioty', []),
        'grupa_info': original_dict.get('grupa_naglowek')
    }

print(transform_dict(dict_original_dict))

import openpyxl
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill
import os

import re
import os
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill

import re
import os
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill


def natural_sort_key(s):
    if s is None:
        return []
    return [int(text) if text.isdigit() else text.lower()
            for text in re.split('([0-9]+)', str(s))]


def generuj(dict_pytania):
    hierarchia = dict_pytania.get('hierarchia', {})
    podmioty = dict_pytania.get('podmioty', [])

    # 1. STYLE
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Kolory teł
    fill_header_row = PatternFill(start_color='365f91', end_color='365f91', fill_type='solid')  # Granatowy (Nagłówek)
    fill_bl = PatternFill(start_color='8db4e2', end_color='8db4e2', fill_type='solid')  # Niebieski
    fill_pb = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')  # Jasny niebieski
    fill_pyt = PatternFill(start_color='d9e3f0', end_color='d9e3f0', fill_type='solid')  # Błękitny
    fill_input = PatternFill(start_color='e2efda', end_color='e2efda', fill_type='solid')  # Zielony (Odpowiedź)


    font_white_bold = Font(bold=True, color="FFFFFF")  # TYLKO Wiersz 3
    font_black_bold = Font(bold=True, color="000000")  # Wszystko poniżej (Nagłówki sekcji)

    for podmiot in podmioty:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Odpowiedzi"

        # Tytuł (Czarny)
        # ws.cell(row=1, column=2, value=f"Ankieta dla: {podmiot['nazwa']} ({podmiot['kod']})").font = font_black_bold

        # --- WIERSZ 3: NAGŁÓWKI KOLUMN (BIAŁY FONT) ---
        headers = ["Nr", "Treść", "Odpowiedź", "Obligatoryjne", "Uzasadnienie"]
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=text)
            cell.font = font_white_bold  # <--- TUTAJ JEDYNY BIAŁY
            cell.fill = fill_header_row
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        row_idx = 4

        # --- WSZYSTKO PONIŻEJ MA WYMUSZONY CZARNY FONT ---

        # PĘTLA 1: OBSZARY
        for obs_kod in sorted(hierarchia.keys(), key=natural_sort_key):
            obs_data = hierarchia[obs_kod]
            nazwa_obs = obs_data['info']['kod']

            # PĘTLA 2: BLOKI
            for bl_kod in sorted(obs_data['bloki'].keys(), key=natural_sort_key):
                bl_data = obs_data['bloki'][bl_kod]
                for c in range(1, 6):
                    cell = ws.cell(row=row_idx, column=c)
                    cell.fill = fill_bl
                    cell.border = thin_border
                    cell.font = font_black_bold  # CZARNY
                ws.cell(row=row_idx, column=1, value=f"{nazwa_obs} {bl_kod}")
                ws.cell(row=row_idx, column=2, value=bl_data['info']['tresc'])
                row_idx += 1

                # PĘTLA 3: PODBLOKI
                for pb_kod in sorted(bl_data['podbloki'].keys(), key=natural_sort_key):
                    pb_data = bl_data['podbloki'][pb_kod]
                    for c in range(1, 6):
                        cell = ws.cell(row=row_idx, column=c)
                        cell.fill = fill_pb
                        cell.border = thin_border
                        cell.font = font_black_bold  # CZARNY
                    ws.cell(row=row_idx, column=1, value=f"{nazwa_obs} {pb_kod}")
                    ws.cell(row=row_idx, column=2, value=pb_data['info']['tresc']).alignment = Alignment(indent=0)
                    row_idx += 1

                    # PĘTLA 4: PYTANIA
                    for py_kod in sorted(pb_data['pytania'].keys(), key=natural_sort_key):
                        py_data = pb_data['pytania'][py_kod]
                        for c in range(1, 6):
                            cell = ws.cell(row=row_idx, column=c)
                            cell.fill = fill_pyt
                            cell.border = thin_border
                            cell.font = font_black_bold  # CZARNY
                        ws.cell(row=row_idx, column=1, value=f"{nazwa_obs} {py_kod}")
                        ws.cell(row=row_idx, column=2, value=py_data['info']['tresc']).alignment = Alignment(indent=0)
                        row_idx += 1

                        # PĘTLA 5: PODZAPYTANIA
                        sorted_pz = sorted(py_data['podzapytania'], key=lambda x: natural_sort_key(x['kod']))
                        for pz in sorted_pz:
                            for c in range(1, 6):
                                cell = ws.cell(row=row_idx, column=c)

                                cell.font = font_black_bold  # CZARNY
                                cell.border = thin_border
                                if c in [3,4, 5]:  # Kolumny do wpisywania
                                    cell.fill = fill_input

                            ws.cell(row=row_idx, column=1, value=f"{nazwa_obs} {pz['kod']}")
                            ws.cell(row=row_idx, column=2, value=pz['tresc']).alignment = Alignment(indent=0)
                            ws.cell(row=row_idx, column=4,
                                    value="TAK" if pz['obligatoryjne'] else "NIE").alignment = Alignment(
                                horizontal='center')
                            row_idx += 1

        # Szerokości kolumn
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 75
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 35
        ws.freeze_panes = "A4"
        # --- NOWA ZAKŁADKA: INFORMACJE ---
        ws_info = wb.create_sheet(title="Informacje")

        ws_info.cell(row=1, column=1, value="Nazwa podmiotu:").font = font_black_bold
        ws_info.cell(row=1, column=2, value=podmiot['nazwa'])

        ws_info.cell(row=2, column=1, value="Kod podmiotu:").font = font_black_bold
        ws_info.cell(row=2, column=2, value=podmiot['kod'])

        ws_info.cell(row=3, column=1, value="Osoba odpowiedzialna za ankietę:").font = font_black_bold
        ws_info.cell(row=3, column=2, value="")  # <- tutaj użytkownik wpisze ręcznie
        ws_info.column_dimensions['A'].width = 50
        ws_info.column_dimensions['B'].width = 75

        output_dir = os.path.join(".", "fossa2", "generated_reports")
        if not os.path.exists(output_dir): os.makedirs(output_dir)
        filename = f"{datetime.datetime.now().year}_{podmiot['kod']}.xlsx"
        wb.save(os.path.join(output_dir, filename))